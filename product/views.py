from decimal import Decimal
from urllib.parse import quote
import io, os, json, qrcode
import matplotlib.pyplot as plt
import pandas as pd

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, FileResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Q
from django.utils.timezone import now
from datetime import timedelta
from django.conf import settings

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart

from .models import Product, Order, OrderItem, Customer, VerificationLog, Shelf
from payment.models import Payment, StockDeductionLog


# ------------------------
# Helpers
# ------------------------
def is_cashier_or_owner(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def get_client_ip(request):
    """Helper: Extract client IP."""
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0]
    return request.META.get("REMOTE_ADDR")


# ------------------------
# AJAX: Cart Count
# ------------------------
def cart_count_ajax(request):
    """Return cart item count as JSON for AJAX updates."""
    cart = request.session.get("cart", {})
    count = sum(item.get("quantity", 1) if isinstance(item, dict) else 1 for item in cart.values())
    return JsonResponse({"count": count})


# ------------------------
# Product CRUD
# ------------------------
@login_required
@user_passes_test(is_cashier_or_owner)
def product_list_admin(request):
    products = Product.objects.select_related("shelf").all()
    return render(request, "product/product_list_admin.html", {"products": products})


@login_required
@user_passes_test(is_cashier_or_owner)
def product_create(request):
    if request.method == "POST":
        shelf_id = request.POST.get("shelf")
        shelf = Shelf.objects.get(id=shelf_id) if shelf_id else None
        Product.objects.create(
            name=request.POST.get("name"),
            price=request.POST.get("price"),
            stock=request.POST.get("stock"),
            barcode=request.POST.get("barcode"),
            shelf=shelf,
        )
        messages.success(request, "Product added successfully.")
        return redirect("product:product_list_admin")

    shelves = Shelf.objects.all()
    return render(request, "product/product_form.html", {"shelves": shelves})


@login_required
@user_passes_test(is_cashier_or_owner)
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product.name = request.POST.get("name")
        product.price = request.POST.get("price")
        product.stock = request.POST.get("stock")
        product.barcode = request.POST.get("barcode")
        shelf_id = request.POST.get("shelf")
        product.shelf = Shelf.objects.get(id=shelf_id) if shelf_id else None
        product.save()
        messages.success(request, "Product updated successfully.")
        return redirect("product:product_list_admin")

    shelves = Shelf.objects.all()
    return render(request, "product/product_form.html", {"product": product, "shelves": shelves})


@login_required
@user_passes_test(is_cashier_or_owner)
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.delete()
    messages.success(request, "Product deleted successfully.")
    return redirect("product:product_list_admin")


# ------------------------
# Public Shopping
# ------------------------
def product_list(request):
    """Customer-facing product list with search and filter."""
    query = request.GET.get("q")
    stock_filter = request.GET.get("stock_filter")
    
    products = Product.objects.all()
    
    if query:
        products = products.filter(
            Q(name__icontains=query) | 
            Q(barcode__icontains=query) |
            Q(description__icontains=query)
        )
    
    if stock_filter == "in_stock":
        products = products.filter(stock__gt=10)
    elif stock_filter == "low_stock":
        products = products.filter(stock__lte=10, stock__gt=0)
    elif stock_filter == "out_of_stock":
        products = products.filter(stock=0)
    
    return render(request, "product/product_list.html", {"products": products})


# ------------------------
# Cart Views
# ------------------------
def cart_view(request):
    cart = request.session.get("cart", {})
    items, total = [], Decimal("0.00")

    for product_id, item_data in cart.items():
        product = get_object_or_404(Product, id=product_id)
        quantity = item_data.get("quantity") if isinstance(item_data, dict) else int(item_data)
        subtotal = product.price * quantity
        items.append({"product": product, "quantity": quantity, "subtotal": subtotal})
        total += subtotal

    return render(request, "product/cart.html", {"items": items, "total": total})


def add_to_cart(request, pk):
    """Add product to cart and return JSON response for AJAX."""
    product = get_object_or_404(Product, pk=pk)
    cart = request.session.get("cart", {})
    pk_str = str(pk)

    if pk_str in cart:
        if isinstance(cart[pk_str], dict):
            cart[pk_str]["quantity"] += 1
        else:
            cart[pk_str] = {"name": product.name, "price": float(product.price), "quantity": int(cart[pk_str]) + 1}
    else:
        cart[pk_str] = {"name": product.name, "price": float(product.price), "quantity": 1}

    request.session["cart"] = cart
    request.session.modified = True

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        quantity = cart[pk_str]["quantity"]
        return JsonResponse(
            {
                "success": True,
                "message": f"{product.name} added to cart",
                "product_name": product.name,
                "cart_count": sum(
                    item.get("quantity", 1) if isinstance(item, dict) else int(item)
                    for item in cart.values()
                ),
                "item_quantity": quantity,
            }
        )

    messages.success(request, f"{product.name} added to cart.")
    return redirect("product:product_list")


def checkout(request):
    """Create an order from session cart and trigger payment redirect."""
    cart = request.session.get("cart", {})
    existing_order_id = request.GET.get("order_id")

    if not cart and existing_order_id:
        order = get_object_or_404(Order, id=existing_order_id)
        items = [
            {
                "product": item.product,
                "quantity": item.quantity,
                "subtotal": item.subtotal(),
            }
            for item in order.items.select_related("product").all()
        ]
        return render(
            request,
            "product/checkout.html",
            {"items": items, "total": order.total_price, "order": order},
        )

    if not cart:
        messages.warning(request, "Your cart is empty.")
        return redirect("product:cart")

    items, total = [], Decimal("0.00")
    for product_id, item_data in cart.items():
        product = get_object_or_404(Product, id=product_id)
        quantity = item_data.get("quantity") if isinstance(item_data, dict) else int(item_data)
        subtotal = product.price * quantity
        items.append({"product": product, "quantity": quantity, "subtotal": subtotal})
        total += subtotal

    order = None
    if request.method == "POST":
        raw_phone = (request.POST.get("phone_number") or "").strip()
        phone_digits = "".join(ch for ch in raw_phone if ch.isdigit())
        if phone_digits.startswith("0"):
            phone_digits = "254" + phone_digits[1:]
        elif phone_digits.startswith("7") and len(phone_digits) == 9:
            phone_digits = "254" + phone_digits

        if not phone_digits.startswith("254") or len(phone_digits) != 12:
            messages.error(request, "Enter a valid Safaricom number (e.g. 07XXXXXXXX).")
            return render(request, "product/checkout.html", {"items": items, "total": total, "order": order})

        action = request.POST.get("action")
        customer, _ = Customer.objects.get_or_create(phone_number=phone_digits)
        order = Order.objects.create(customer=customer, status="PENDING", total_price=total)

        for product_id, qty in cart.items():
            product = get_object_or_404(Product, id=product_id)
            quantity = qty["quantity"] if isinstance(qty, dict) else int(qty)
            OrderItem.objects.create(order=order, product=product, quantity=quantity, price=product.price)

        request.session["cart"] = {}
        request.session.modified = True

        if action == "pay_now":
            return render(
                request,
                "payment/payment_redirect.html",
                {"order": order, "phone_number": phone_digits},
            )
        return redirect("product:receipt", order_id=order.id)

    return render(request, "product/checkout.html", {"items": items, "total": total, "order": order})



@login_required
@user_passes_test(is_cashier_or_owner)
def mark_order_paid(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if order.status != "PAID":
        order.status = "PAID"
        order.save()
        apply_stock_deduction(order)
        messages.success(request, f"Order #{order.id} marked as PAID and stock updated.")
    else:
        messages.warning(request, f"Order #{order.id} is already PAID.")
    return redirect("product:dashboard")


def apply_stock_deduction(order):
    """Reduce stock for each item when paid (log once)."""
    for item in order.items.all():
        if item.product.stock >= item.quantity:
            item.product.stock -= item.quantity
            item.product.save()
            StockDeductionLog.objects.create(product=item.product, order=order, quantity=item.quantity, action="DEDUCT")


# ------------------------
# Dashboard + Reports
# ------------------------
@login_required
@user_passes_test(is_cashier_or_owner)
def dashboard(request):
    filter_option = request.GET.get("filter", "month")
    today = now().date()

    start_date = today if filter_option == "today" else today - timedelta(days=7 if filter_option == "week" else 30)

    paid_orders = Order.objects.filter(status="PAID", created_at__date__gte=start_date)
    pending_orders = Order.objects.filter(status="PENDING", created_at__date__gte=start_date)
    cancelled_orders = Order.objects.filter(status="CANCELLED", created_at__date__gte=start_date)
    failed_orders = Order.objects.filter(status="FAILED", created_at__date__gte=start_date)

    total_sales = paid_orders.aggregate(total=Sum("total_price"))["total"] or 0
    total_orders = paid_orders.count() + pending_orders.count() + cancelled_orders.count() + failed_orders.count()
    total_products = Product.objects.count()
    total_stock = Product.objects.aggregate(total=Sum("stock"))["total"] or 0

    refunds = Payment.objects.filter(status="REFUNDED", transaction_date__date__gte=start_date)
    refund_total = refunds.aggregate(total=Sum("amount"))["total"] or 0
    refund_ratio = (refund_total / total_sales * 100) if total_sales > 0 else 0

    uncollected_orders = Order.objects.filter(status="PAID").exclude(payments__isnull=True)

    low_stock_products = Product.objects.filter(stock__lte=5)
    recent_orders = Order.objects.filter(created_at__date__gte=start_date).order_by("-created_at")[:10]

    order_status_data = {
        "Paid": paid_orders.count(),
        "Pending": pending_orders.count(),
        "Cancelled": cancelled_orders.count(),
        "Failed": failed_orders.count(),
        "Refunded": refunds.count(),
    }

    sales_trends = [{"date": (today - timedelta(days=i)).strftime("%Y-%m-%d"),
                     "sales": Order.objects.filter(status="PAID", created_at__date=today - timedelta(days=i))
                     .aggregate(total=Sum("total_price"))["total"] or 0} for i in range(7)][::-1]

    refunds_trends = [{"date": (today - timedelta(days=i)).strftime("%Y-%m-%d"),
                       "refunds": Payment.objects.filter(status="REFUNDED", transaction_date__date=today - timedelta(days=i)).count()} for i in range(7)][::-1]

    top_products = (OrderItem.objects.filter(order__status="PAID", order__created_at__date__gte=start_date)
                    .values("product__name", "product__shelf__name")
                    .annotate(total_sold=Sum("quantity")).order_by("-total_sold")[:5])
    top_products_data = {f"{item['product__name']} ({item['product__shelf__name']})": item["total_sold"] for item in top_products}

    refund_vs_sales_data = {"Sales": float(total_sales - refund_total), "Refunds": float(refund_total)}

    return render(request, "product/dashboard.html", {
        "total_sales": total_sales,
        "refund_total": refund_total,
        "refund_ratio": refund_ratio,
        "refund_vs_sales_data": json.dumps(refund_vs_sales_data),
        "total_orders": total_orders,
        "total_products": total_products,
        "total_stock": total_stock,
        "paid_count": paid_orders.count(),
        "pending_count": pending_orders.count(),
        "cancelled_count": cancelled_orders.count(),
        "failed_count": failed_orders.count(),
        "refund_count": refunds.count(),
        "low_stock_products": low_stock_products,
        "recent_orders": recent_orders,
        "uncollected_orders": uncollected_orders,
        "order_status_data": json.dumps(order_status_data),
        "sales_trends": json.dumps(sales_trends),
        "refunds_trends": json.dumps(refunds_trends),
        "top_products_data": json.dumps(top_products_data),
        "filter_option": filter_option,
    })

# ------------------------
# Receipts + Verification
# ------------------------
from django.shortcuts import render, get_object_or_404
from payment.models import Payment  # import your Payment model

def receipt(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    receipt_no = f"RCPT-{order.id:05d}"
    transaction_date = order.created_at.strftime("%Y-%m-%d %H:%M:%S")
    phone_number = order.customer.phone_number if order.customer else "N/A"

    # ✅ Get latest payment if exists
    payment = Payment.objects.filter(order=order).last()
    verify_path = request.build_absolute_uri(f"/order/{order.id}/verify/")

    whatsapp_url = None
    if order.customer and order.customer.phone_number and payment and payment.mpesa_receipt_no:
        phone_digits = "".join(ch for ch in order.customer.phone_number if ch.isdigit())
        if phone_digits.startswith("254"):
            wa_phone = phone_digits
        elif phone_digits.startswith("0"):
            wa_phone = f"254{phone_digits[1:]}"
        elif phone_digits.startswith("7") and len(phone_digits) == 9:
            wa_phone = f"254{phone_digits}"
        else:
            wa_phone = phone_digits

        wa_text = (
            f"Order #{order.id} has been PAID. "
            f"M-Pesa receipt: {payment.mpesa_receipt_no}. "
            f"Total: KES {order.total_price:.2f}. "
            f"Verify: {verify_path}"
        )
        whatsapp_url = f"https://wa.me/{wa_phone}?text={quote(wa_text)}"

    return render(request, "product/receipt.html", {
        "order": order,
        "receipt_no": receipt_no,
        "transaction_date": transaction_date,
        "phone_number": phone_number,
        "payment": payment,   # ✅ pass payment
        "verify_path": verify_path,
        "whatsapp_url": whatsapp_url,
    })


def receipt_pdf(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)

    # Header
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, f"Receipt - Order #{order.id}")

    p.setFont("Helvetica", 12)
    p.drawString(100, 780, f"Date: {order.created_at.strftime('%Y-%m-%d %H:%M:%S')}")
    p.drawString(100, 765, f"Customer: {order.customer.phone_number if order.customer else 'N/A'}")
    p.drawString(100, 750, f"Status: {order.status}")

    # Table header
    y = 720
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, y, "Product")
    p.drawString(250, y, "Qty")
    p.drawString(300, y, "Price")
    p.drawString(380, y, "Subtotal")

    # Table rows
    p.setFont("Helvetica", 12)
    y -= 20
    for item in order.items.all():
        p.drawString(100, y, item.product.name)
        p.drawString(250, y, str(item.quantity))
        p.drawString(300, y, str(item.price))
        p.drawString(380, y, str(item.subtotal))
        y -= 20

    # Total
    y -= 20
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, y, f"Total: {order.total_price}")

    p.showPage()
    p.save()
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename=f"receipt_{order.id}.pdf")


def verify_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    VerificationLog.objects.create(order=order, ip_address=get_client_ip(request))
    return render(request, "product/verify_order.html", {"order": order})


# ------------------------
# Payment Status
# ------------------------
from django.http import JsonResponse

def check_payment_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    payment = Payment.objects.filter(order=order).last()

    if payment and payment.status == "PAID":
        order.status = "PAID"
        order.save(update_fields=["status"])
        return JsonResponse({"status": "PAID"})

    return JsonResponse({"status": order.status})


# ------------------------
# Export Reports
# ------------------------
@login_required
@user_passes_test(is_cashier_or_owner)
def export_excel(request):
    today = now().date()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header
    ws.append(["Date", "Order ID", "Customer", "Total Price", "Status"])

    orders = Order.objects.filter(created_at__date=today)
    for o in orders:
        ws.append([o.created_at.strftime("%Y-%m-%d"), o.id, o.customer.phone_number, float(o.total_price), o.status])

    # Add Sales Trend Chart
    chart = LineChart()
    data = Reference(ws, min_col=4, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=False)
    chart.title = "Daily Sales Trend"
    ws.add_chart(chart, "G2")

    # Add Refund Chart
    refund_ws = wb.create_sheet(title="Refunds")
    refund_ws.append(["Date", "Refund Amount"])

    refunds = Payment.objects.filter(status="REFUNDED", transaction_date__date=today)
    for r in refunds:
        refund_ws.append([r.transaction_date.strftime("%Y-%m-%d"), float(r.amount)])

    refund_chart = BarChart()
    refund_data = Reference(refund_ws, min_col=2, min_row=2, max_row=refund_ws.max_row)
    refund_chart.add_data(refund_data, titles_from_data=False)
    refund_chart.title = "Refunds Trend"
    refund_ws.add_chart(refund_chart, "D2")

    # Save to response
    filename = f"sales_report_{today}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_cashier_or_owner)
def sales_graph(request):
    today = now().date()
    orders = (Order.objects.filter(status="PAID", created_at__date__gte=today - timedelta(days=30))
              .values("created_at__date").annotate(total=Sum("total_price")).order_by("created_at__date"))

    df = pd.DataFrame(list(orders))
    if df.empty:
        df = pd.DataFrame({"created_at__date": [today], "total": [0]})

    plt.figure(figsize=(8, 4))
    plt.plot(df["created_at__date"], df["total"], marker="o")
    plt.title("Sales Over Last 30 Days")
    plt.xlabel("Date")
    plt.ylabel("Sales")
    plt.xticks(rotation=45)

    img_path = os.path.join(settings.MEDIA_ROOT, "sales_graph.png")
    plt.tight_layout()
    plt.savefig(img_path)
    plt.close()

    with open(img_path, "rb") as f:
        return HttpResponse(f.read(), content_type="image/png")


def remove_from_cart(request, pk):
    """Remove one product from cart."""
    cart = request.session.get("cart", {})
    product_id = str(pk)

    if product_id in cart:
        del cart[product_id]
        request.session["cart"] = cart
        request.session.modified = True

    return redirect("product:cart")


def update_cart(request, pk):
    """Update quantity of a product in the cart."""
    if request.method == "POST":
        try:
            new_qty = int(request.POST.get("quantity", 1))
        except ValueError:
            new_qty = 1

        cart = request.session.get("cart", {})
        product_id = str(pk)

        if new_qty > 0 and product_id in cart:
            cart[product_id]["quantity"] = new_qty
        else:
            cart.pop(product_id, None)

        request.session["cart"] = cart
        request.session.modified = True

    return redirect("product:cart")


def clear_cart(request):
    """Clear entire cart."""
    request.session["cart"] = {}
    request.session.modified = True
    return redirect("product:cart")


# -----------------------
# ORDER RECEIPT
# -----------------------
def receipt_view(request, order_id):
    """Display receipt for an order."""
    return redirect("product:receipt", order_id=order_id)


def send_receipt_sms(request, order_id):
    """Send a receipt verification link to the order customer's phone."""
    order = get_object_or_404(Order, id=order_id)

    if not order.customer or not order.customer.phone_number:
        messages.error(request, "No customer phone number is available for this order.")
        return redirect("product:receipt", order_id=order.id)

    from payment.sms import send_sms

    receipt_link = request.build_absolute_uri(f"/receipt/{order.id}/")
    message = (
        f"Your receipt for Order #{order.id}. "
        f"Status: {order.status}. "
        f"View receipt: {receipt_link}"
    )
    success, details = send_sms(order.customer.phone_number, message)

    if success:
        messages.success(request, "Receipt SMS sent successfully.")
    else:
        messages.error(request, f"Failed to send receipt SMS: {details}")

    return redirect("product:receipt", order_id=order.id)
